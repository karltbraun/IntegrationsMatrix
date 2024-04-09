from openpyxl import load_workbook

# Open the file /Users/karlbraun/Documents/DEV-L/KTB/Misc/your_file.xlsx
wb = load_workbook("/Users/karlbraun/Documents/DEV-L/KTB/Misc/your_file.xlsx")
ws = wb.active

comment = ws["A1"].comment
print(comment.text)
