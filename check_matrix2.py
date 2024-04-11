from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Tuple, Dict
import os


""" check_matrix - Compare values in a matrix with their transposed positions.`
"""


# TEST = True
TEST = False

if TEST:
    FILENAME_IN: str = "/Users/karlbraun/Documents/DEV-L/KTB/Misc/test_file.xlsx"
    TABLE_NAME: str = "A1:F6"
else:
    FILENAME_IN: str = (
        "/Users/karlbraun/Documents/DEV-L/KTB/Misc/MDR_Vendor_Comparisons.xlsx"
    )
    TABLE_NAME: str = "B2:AL35"


# ################### Excel Table to Matrix ###################


def excel_table_to_matrix(table: Worksheet) -> List[List[str]]:
    matrix = []
    for row in table:
        row_values = [
            cell.value.strip().upper() if cell.value is not None else None
            for cell in row
        ]
        matrix.append(row_values)
    return matrix


# ################### Transform Data ###################


def transform_data(
    data: List[List[str]],
) -> Tuple[Dict[str, Dict[str, str]], Dict[str, Dict[str, str]]]:

    row_dict = {}
    col_dict = {}

    row_headers = data[0][1:]
    col_headers = [row[0] for row in data[1:]]

    for i, row in enumerate(data[1:], start=1):
        row_header = row[0]
        row_dict[row_header] = {
            col_header: row[j] for j, col_header in enumerate(row_headers, start=1)
        }

    for j, col_header in enumerate(row_headers, start=1):
        col_dict[col_header] = {row[0]: row[j] for row in data[1:]}

    return row_dict, col_dict


# ################### Check for None ###################


def check_for_none(value: str) -> str:
    if value is None:
        return "<None>"
    else:
        return value


# ################### Compare Values 2 ###################


def compare_values2(
    matrix: List[List[str]],
) -> Tuple[List[Tuple[str, str, str]], List[Tuple[str, str, str, str, str, str]]]:

    print("ZZZZZZZZZZZZZZZZZZZ compare_values2 ZZZZZZZZZZZZZZZZZZZ")
    # Extract column headers and row headers
    col_headers = matrix[0][1:]
    row_headers = [row[0] for row in matrix[1:]]

    output_matrix = []
    error_report = []

    # Create a set to keep track of processed pairs
    processed_pairs = set()

    num_rows = len(matrix)
    num_cols = len(matrix[0])

    for i in range(1, num_rows):  # iterate over the data cells in the matrix
        for j in range(1, num_cols):
            h1 = row_headers[i - 1]
            h2 = col_headers[j - 1]
            v = matrix[i][j]
            print(f"checking h1={h1} h2={h2} v={v}")

            # Check if the pair (h1, h2) or (h2, h1) has already been processed
            if (h1, h2) not in processed_pairs and (h2, h1) not in processed_pairs:
                # Check if (h2, h1) is a valid reference
                if h2 in row_headers and h1 in col_headers:
                    h2_index = row_headers.index(h2) + 1
                    h1_index = col_headers.index(h1) + 1
                    v2 = matrix[h2_index][h1_index]

                    # Compare values for (h1, h2) and (h2, h1)
                    if v == v2:
                        output_matrix.append((h1, h2, v))
                    else:
                        error_report.append((h1, h2, v, h2, h1, v2))
                else:
                    output_matrix.append((h1, h2, v))

                processed_pairs.add((h1, h2))
                processed_pairs.add((h2, h1))

    output_matrix.sort(key=lambda x: x[0])
    return output_matrix, error_report


# ################### Print Report ###################


def print_report(
    results: List[Tuple[Tuple[str, str], Tuple[str, str], str, str, str]]
) -> None:
    """
    Print the comparison results in a formatted report.
    """
    return

    for result in results:
        print(f"Comparison: {result[0]} vs {result[1]}")
        print(
            f"Is Equal: {result[2]:<20}Value: {result[3]:<20}Transposed Value: {result[4]}"
        )
        print()


# ################### Load Matrix ###################


def load_matrix(filename: str, table_name: str) -> List[List[str]]:

    matrix: List[List[str]] = []

    if not os.path.exists(filename):
        print(f"File '{filename}' does not exist.")
    elif not os.access(filename, os.R_OK):
        print(f"Cannot read file '{filename}'.")
    else:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        table = ws[table_name]
        matrix = excel_table_to_matrix(table)

    return matrix


# ################## print dictionary ##################


def print_dict(dictionary: Dict[str, Dict[str, str]], title: str) -> None:
    print(f"########## {title} ##########")
    for key, value in dictionary.items():
        print(f"{key}:")
        for k, v in value.items():
            print(f"  {k}: {v}")
        print()
    print()


# ################### MAIN ###################


def main():
    # read in the workbook and get the first worksheet
    if (input_data := load_matrix(FILENAME_IN, TABLE_NAME)) is None:
        return 1

    # Transform data into dictionaries
    row_dict, col_dict = transform_data(input_data)

    # Create a unique set of all values in row_dict and col_dict
    all_vendors = set()
    all_vendors.update(*row_dict.values())
    all_vendors.update(*col_dict.values())

    print("--------------------- Number of Vendors ---------------------")
    print(f"Number of Vendors: {len(all_vendors)}")
    print(all_vendors)

    # Print the two dictionaries
    print_dict(row_dict, "Row Dictionary")
    print_dict(col_dict, "Column Dictionary")

    # Compare values and generate results
    output_matrix, error_report = compare_values2(input_data)

    print(
        f"===================== Comparison Results ({len(output_matrix)}) ====================="
    )
    print(f"Items successfully compared: {len(output_matrix)}")
    for item in output_matrix:
        print(item)
    print()

    # Print the error report
    print("##################### Error Report #####################")
    print_report(error_report)

    print("##################### Total Integrations #####################")
    vendor_integrations = []
    for vendor in all_vendors:
        count = sum(
            1
            for item in output_matrix
            if (item[0] == vendor or item[1] == vendor) and item[2] == "YES"
        )
        vendor_integrations.append((vendor, count))

    vendor_integrations.sort(key=lambda x: x[1], reverse=True)
    for vendor, count in vendor_integrations:
        print(f"{vendor} has {count} valid integrations with other vendors")

    return 0


if __name__ == "__main__":
    exit(main())
