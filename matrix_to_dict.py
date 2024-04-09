from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from check_headers import check_headers
import sys

FILENAME_IN_TEST: str = "/Users/karlbraun/Documents/DEV-L/KTB/Misc/test_file.xlsx"
FILENAME_IN_REAL: str = (
    "/Users/karlbraun/Documents/DEV-L/KTB/Misc/MDR_Vendor_Comparisons.xlsx"
)
FILENAME_IN: str = FILENAME_IN_TEST
TABLE_NAME: str = "B2:AL35"


def excel_table_to_matrix(table: Worksheet) -> List[List[str]]:
    matrix = []
    for row in table:
        row_values = [
            cell.value.strip().upper() if cell.value is not None else None
            for cell in row
        ]
        matrix.append(row_values)
    return matrix


def convert_matrix_to_dictionary(input_matrix):
    # Extract column headers (excluding the first element)
    column_headers = input_matrix[0][1:]

    # Extract row headers (excluding the first row)
    row_headers = [row[0] for row in input_matrix[1:]]

    # Create the first dictionary (row-based)
    row_dict = {}
    for row in input_matrix[1:]:
        row_header = row[0]
        row_values = row[1:]
        row_dict[row_header] = dict(zip(column_headers, row_values))

    # Create the second dictionary (column-based)
    col_dict = {}
    for col_index, col_header in enumerate(column_headers, start=1):
        col_values = [row[col_index] for row in input_matrix[1:]]
        col_dict[col_header] = dict(zip(row_headers, col_values))

    return row_dict, col_dict


if __name__ == "__main__":

    # read in the workbook and get the first worksheet
    wb = load_workbook(FILENAME_IN)
    ws = wb.worksheets[0]
    table = ws[TABLE_NAME]

    # convert the range to a 2D matrix
    matrix = excel_table_to_matrix(table)
    print("********** Input Matrix **********")
    for row in matrix:
        print(row)
        print()
    print()

    # convert the matrix to a pair of dictionaries (row-based and column-based)
    row_dict, col_dict = convert_matrix_to_dictionary(matrix)

    print("********** Row Dictionary **********")
    for key, value in row_dict.items():
        print(key, ":", value)
        print()
    print()

    print("********** Column Dictionary **********")
    for key, value in col_dict.items():
        print(key, ":", value)
        print()
    print()

    sys.exit(0)
