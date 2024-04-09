from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
import sys

FILENAME_IN: str = "/Users/karlbraun/Documents/DEV-L/KTB/Misc/your_file.xlsx"
TABLE_NAME: str = "tbl_integrations"


def check_headers(input_matrix):
    # Extract column headers (1st row of the matrix)
    column_headers = input_matrix[0][1:]

    # Extract row headers (1st column of each row)
    row_headers = [row[0] for row in input_matrix[1:]]

    # Convert both lists to upper case
    # column_headers = [header.upper() for header in column_headers]
    # row_headers = [header.upper() for header in row_headers]
    column_headers = [header for header in column_headers]
    row_headers = [header for header in row_headers]

    # Sort both lists
    column_headers.sort()
    row_headers.sort()

    print("########## Headers ##########")
    print("Column headers:", column_headers)
    print("Row headers:", row_headers)
    print()

    # Find row headers with no matching column header
    unmatched_row_headers = [
        header for header in row_headers if header not in column_headers
    ]

    # Find column headers with no matching row header
    unmatched_column_headers = [
        header for header in column_headers if header not in row_headers
    ]

    # Print the results
    print("Number of column headers:", len(column_headers))
    print("Number of row headers:", len(row_headers))
    print("Row headers with no matching column header:", unmatched_row_headers)
    print("Column headers with no matching row header:", unmatched_column_headers)


# ########################### MAIN #########################


if __name__ == "__main__":
    matrix = []
    """
    matrix.append(
        ["Vendors", "Vendor 1", "Vendor 2", "Vendor 3", "Vendor 4", "Vendor 5"]
    )
    matrix.append(["Vendor 1", "Foo", "Bar", "Baz", "Qux", "Zuuz"])
    matrix.append(["Vendor 2", "Foo", "Bar", "Baz", "Qux", "Zuuz"])
    matrix.append(["Vendor 3", "Foo", "Bar", "Baz", "Qux", "Zuuz"])
    matrix.append(["Vendor 4", "Foo", "Bar", "Baz", "Qux", "Zuuz"])
    matrix.append(["Vendor 5", "Foo", "Bar", "Baz", "Qux", "Zuuz"])
    """
    wb = load_workbook(FILENAME_IN)
    ws = wb.active
    matrix = ws["A2:R19"]
    # matrix = ws["tbl_integrations"]

    print("********** Input Matrix **********")
    for row in matrix:
        print(row)
    print()

    check_headers(matrix)

    sys.exit(0)
