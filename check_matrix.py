""" check_matrix.py - Read in an Excel file with a matrix of integrations and
    limited integrations between vendors and output the results.

    Current version has hard coded file names in setup section
"""

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Tuple, Dict
import os
import datetime

# ################### setup ###################

# PRINT_SIMPLE = True  # Set to True to print a simple text output
PRINT_SIMPLE = False  # Set to False to create an Excel file with the results

FILENAME_OUT = "./Data/Security and IT Service Provider Integrations 20240411.xlsx"

# TEST = True  # Set to True to test with a small test file
TEST = False  # Set to False to test with the full file

if TEST:
    FILENAME_IN: str = "/Users/karlbraun/Documents/DEV-L/KTB/Misc/test_file.xlsx"
    TABLE_NAME: str = "A1:F6"
else:
    FILENAME_IN: str = (
        "/Users/karlbraun/Documents/DEV-L/KTB/Misc/MDR_Vendor_Comparisons.xlsx"
    )
    TABLE_NAME: str = "B2:AL35"

PRINT_INPUT_MATRIX = False

# ################### Vendor class ###################


""" Vendor class - Class to hold information about a vendor and their integrations.
    We use this to list the other vendors with which a vendor has integrations 
"""


class Vendor:
    def __init__(self, name):
        self.name = name
        self.integrations = set()
        self.limited = set()

    def add_integration(self, integration):
        self.integrations.add(integration)

    def add_limited(self, integration):
        self.limited.add(integration)

    @property
    def integrations_count(self):
        return len(self.integrations)

    @property
    def limited_count(self):
        return len(self.limited)


# ################### Excel Table to Matrix ###################


from typing import List
from openpyxl.worksheet.worksheet import Worksheet


def excel_table_to_matrix(table: Worksheet) -> List[List[str]]:
    """
    Converts an Excel table into a matrix.

    Args:
        table (Worksheet): The Excel table to convert.

    Returns:
        List[List[str]]: The matrix representation of the Excel table.
    """
    matrix = []
    for row in table:
        row_values = [
            cell.value.strip().upper() if cell.value is not None else None
            for cell in row
        ]
        matrix.append(row_values)
    return matrix


# ################### process_matrix ###################


def process_matrix(matrix: List[List[str]]) -> Dict[str, Vendor]:
    """process_matrix - Process the matrix and return a dictionary of Vendor objects."""

    row_headers = [row[0] for row in matrix[1:]]
    col_headers = matrix[0][1:]
    all_headers = sorted(set(row_headers + col_headers))

    vendors = {header: Vendor(header) for header in all_headers}

    def check_integration(vendor1, vendor2):
        if vendor1 in row_headers and vendor2 in col_headers:
            i = row_headers.index(vendor1) + 1
            j = col_headers.index(vendor2) + 1
            value = matrix[i][j]

            if value == "YES":
                vendors[vendor1].add_integration(vendor2)
                vendors[vendor2].add_integration(vendor1)
                return True
            elif value == "LIMITED":
                vendors[vendor1].add_limited(vendor2)
                vendors[vendor2].add_limited(vendor1)
                return True
        return False

    for vendor1, vendor2 in (
        (v1, v2) for v1 in all_headers for v2 in all_headers if v1 != v2
    ):
        check_integration(vendor1, vendor2) or check_integration(vendor2, vendor1)

    for vendor in vendors.values():
        vendor.integrations = sorted(vendor.integrations)
        vendor.limited = sorted(vendor.limited)

    return vendors


# ################### Load Matrix ###################


def load_matrix(filename: str, table_name: str) -> List[List[str]]:
    """load_matrix - Load a matrix from an Excel file."""

    matrix: List[List[str]] = None

    if not os.path.exists(filename):
        print(f"File '{filename}' does not exist.")
    elif not os.access(filename, os.R_OK):
        print(f"Cannot read file '{filename}'.")
    else:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        table = ws[table_name]
        matrix = excel_table_to_matrix(table)

    return matrix


# ################### get_matrix_input  ###################


def get_matrix_input() -> List[List[str]]:
    """get_matrix_input - dummy routine to mock up getting input from
    hard coded file names and table range.  Ideally this would be
    something more dynamic, getting in put file from command line
    or through dialoge witht the user.
    """
    filename = FILENAME_IN  # defined at the top of the file
    table_name = TABLE_NAME  # defined at the top of the file
    matrix = load_matrix(filename, table_name)
    return matrix


# ################### Create XLSX  ###################


def create_xlsx(vendors: Dict[str, Vendor]) -> None:
    filename = FILENAME_OUT  # defined at the top of the file

    if os.path.exists(filename):
        os.remove(filename)

    wb = Workbook()
    wb.save(filename)

    ws = wb.active
    ws["A1"] = "Date Created"
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    ws["B1"] = current_date

    ws["A2"] = "Vendor"
    ws["B2"] = "Integrations"
    ws["C2"] = "Limited"

    row = 3
    for vendor_name, vendor_data in vendors.items():
        max_len = max(len(vendor_data.integrations), len(vendor_data.limited))
        for i in range(max_len):
            ws.cell(row=row, column=1, value=vendor_name)
            ws.cell(
                row=row,
                column=2,
                value=(
                    vendor_data.integrations[i]
                    if i < len(vendor_data.integrations)
                    else ""
                ),
            )
            ws.cell(
                row=row,
                column=3,
                value=vendor_data.limited[i] if i < len(vendor_data.limited) else "",
            )
            row += 1

    wb.save(filename)


# ################### Print Simple  ###################


def print_simple(vendors: Dict[str, Vendor]) -> None:
    """print a simple text output listing each vendor, the list of vendors
    they have integrations with and the list of vendors they have limited
    integrations with.
    """
    current_time = datetime.datetime.now()
    print(f"----------------------- {current_time} -----------------------")
    print("\nIntegration Results:")
    for vendor in vendors.values():
        print(f"Vendor: {vendor.name}")
        print(f"  Integrations: {vendor.integrations_count}\n  {vendor.integrations}")
        print(f"  Limited: {vendor.limited_count}\n  {vendor.limited}")
        print()


# ################### MAIN ###################


def main():
    # read in the workbook and get the first worksheet
    if (input_data := get_matrix_input()) is None:
        return 1

    if PRINT_INPUT_MATRIX:
        print("Input matrix:")
        for row in input_data:
            print(row)
        print()

    vendors = process_matrix(input_data)

    if PRINT_SIMPLE:
        print_simple(vendors)
    else:
        create_xlsx(vendors)


if __name__ == "__main__":
    exit(main())
