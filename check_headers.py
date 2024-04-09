from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
import sys

FILENAME_IN: str = "/Users/karlbraun/Documents/DEV-L/KTB/Misc/your_file.xlsx"
TABLE_NAME: str = "tbl_integrations"


def check_headers(input_matrix):
    # Extract column headers (1st row of the matrix)
    column_headers = input_matrix[0][1:]
    # Extract row headers (1st column of each row)
    row_headers = [row[0] for row in input_matrix[1:]]

    # Convert both lists to upper case
    # column_headers = [header.upper() for header in column_headers]
    # row_headers = [header.upper() for header in row_headers]
    column_headers = [header.upper().strip() for header in column_headers]
    row_headers = [header.upper().strip() for header in row_headers]

    # Sort both lists
    column_headers.sort()
    row_headers.sort()

    print("########## Headers ##########")
    print("Column headers:", column_headers)
    print("Row headers:", row_headers)
    print()

    # Find row headers with no matching column header
    unmatched_row_headers = [
        header for header in row_headers if header not in column_headers
    ]

    # Find column headers with no matching row header
    unmatched_column_headers = [
        header for header in column_headers if header not in row_headers
    ]

    # Print the results
    print("#################### Header Check ####################")
    print("Number of column headers:", len(column_headers))
    print("Number of row headers:", len(row_headers))
    print("Row headers with no matching column header:", unmatched_row_headers)
    print("Column headers with no matching row header:", unmatched_column_headers)

    rv = 1
    if len(unmatched_row_headers) == 0 and len(unmatched_column_headers) == 0:
        rv = 0

    return rv


# ########################### MAIN #########################


if __name__ == "__main__":
    wb = load_workbook(FILENAME_IN)
    ws = wb.worksheets[0]
    table = ws["A2:R19"]

    matrix = []
    for row in table:
        row_values = [cell.value for cell in row]
        matrix.append(row_values)

    print("********** Input Matrix **********")
    for row in matrix:
        print(row)
    print()

    check_headers(matrix)

    sys.exit(0)
